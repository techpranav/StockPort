import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def format_financial_sheet(worksheet, title):
    """Format the financial sheet with proper styling."""
    # Format header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format title
    worksheet['A1'] = title
    worksheet['A1'].font = Font(size=14, bold=True)
    worksheet.merge_cells('A1:E1')
    
    # Format headers
    for cell in worksheet[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Format data cells
    for row in worksheet.iter_rows(min_row=3):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='right')
    
    # Adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

def generate_excel_report(stock_summaries, excel_filename):
    """Generate an Excel file summarizing stock analysis and financial data."""
    # Create summary DataFrame
    df_summary = pd.DataFrame(
        stock_summaries,
        columns=["Serial Number", "Stock Name", "Overview", "Summary", "Decision"]
    )

    # Create Excel writer
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        # Write summary sheet
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format summary sheet
        summary_sheet = writer.sheets['Summary']
        format_financial_sheet(summary_sheet, "Stock Analysis Summary")
        
        # Process each stock's financial data
        for stock_data in stock_summaries:
            if len(stock_data) > 5:  # Check if financial data exists
                stock_name = stock_data[1]
                financials = stock_data[5]  # Financial data is at index 5
                
                # Create sheets for each financial statement
                if 'financials' in financials:
                    # Income Statement
                    if 'income_statement' in financials['financials']:
                        income_df = pd.DataFrame(financials['financials']['income_statement'])
                        income_df.to_excel(writer, sheet_name=f'{stock_name}_Income', index=False)
                        format_financial_sheet(writer.sheets[f'{stock_name}_Income'], f"{stock_name} - Income Statement")
                    
                    # Balance Sheet
                    if 'balance_sheet' in financials['financials']:
                        balance_df = pd.DataFrame(financials['financials']['balance_sheet'])
                        balance_df.to_excel(writer, sheet_name=f'{stock_name}_Balance', index=False)
                        format_financial_sheet(writer.sheets[f'{stock_name}_Balance'], f"{stock_name} - Balance Sheet")
                    
                    # Cash Flow
                    if 'cashflow' in financials['financials']:
                        cash_df = pd.DataFrame(financials['financials']['cashflow'])
                        cash_df.to_excel(writer, sheet_name=f'{stock_name}_CashFlow', index=False)
                        format_financial_sheet(writer.sheets[f'{stock_name}_CashFlow'], f"{stock_name} - Cash Flow")

    return excel_filename
