import requests as requests

from services.data_providers.yahoo_finance.yahoo_finance_service import YahooFinanceService
from ai.chatgpt_api import get_stock_summary, getShortSummary, getOverallShortSummary
from reports.generate_word import generate_word_report
from reports.generate_excel import generate_excel_report
import json
import pandas as pd
import re
import time
from bs4 import BeautifulSoup

import os
from config.constants.StringConstants import input_dir, output_dir
from utils.path_utils import getSymbolOutputDirectory, getOutputDirectory
from utils.google_drive_utils import authenticate_drive, create_drive_folder, upload_file_to_drive
import traceback
import openpyxl
import shutil

STOCK_FILE = f"{input_dir}\\stocks.txt"
COMPLETED_FILE = f"{input_dir}\\completed.txt"
FAILED_FILE = f"{input_dir}\\failed.txt"
EXCEL_FILE = f"{getOutputDirectory()}\\stock_summary.xlsx"


def read_stock_symbols():
    """Reads stock symbols from 'stocks.txt'."""
    if not os.path.exists(STOCK_FILE):
        print(f"{STOCK_FILE} not found!")
        return []
    with open(STOCK_FILE, "r") as f:
        return [line.strip() for line in f if line.strip()]


def update_stock_file(remaining_symbols):
    """Writes the remaining symbols back to 'stocks.txt'."""
    with open(STOCK_FILE, "w") as f:
        for symbol in remaining_symbols:
            f.write(symbol + "\n")


def append_to_completed(symbol):
    """Appends a completed stock symbol to 'completed.txt'."""
    with open(COMPLETED_FILE, "a") as f:
        f.write(symbol + "\n")


def append_to_failed(symbol):
    """Appends a failed stock symbol to 'failed.txt'."""
    with open(FAILED_FILE, "a") as f:
        f.write(symbol + "\n")


def save_stock_data_to_file(stock_symbol, stock_data):
    """Save stock data (JSON) into a file."""
    processed_data = {str(key): value for key, value in stock_data.items() if not isinstance(value, pd.DataFrame)}
    filename = os.path.join(getSymbolOutputDirectory(stock_symbol), f"{stock_symbol}_stock_data.json")
    with open(filename, "w", encoding="utf-8") as file:
        json.dump(processed_data, file, indent=4, ensure_ascii=False)
    print(f"Stock data saved to: {filename}")


def format_financials(financial_dict, title):
    formatted_str = f"\n### {title} ###\n"
    for key, values in financial_dict.items():
        formatted_str += f"{key}: {values}\n"
    return formatted_str


# Function to scrape Net Income (TTM) from Yahoo Finance Statistics tab

def scrape_yahoo_statistics(ticker):
    """Scrapes all key-value pairs from Yahoo Finance Statistics tab."""
    url = f"https://finance.yahoo.com/quote/{ticker}/key-statistics?p={ticker}"
    headers = {"User-Agent": "Mozilla/5.0"}

    statistics_data = {}

    try:
        response = requests.get(url, headers=headers)
        soup = BeautifulSoup(response.text, "html.parser")

        # Find all statistic rows in the table
        for row in soup.find_all("tr"):
            tds = row.find_all("td")  # Get all <td> elements in the row

            if len(tds) == 2:  # Ensure row has exactly 2 columns (Key, Value)
                key = tds[0].text.strip()
                value = tds[1].text.strip()
                statistics_data[key] = value

        return statistics_data

    except Exception as e:
        print(f"Error scraping statistics for {ticker}: {e}")
        return {}

def convert_to_numeric(value):
    """
    Converts financial values like '96.15B', '850M' to numeric form.
    Returns 0.0 if the conversion fails.
    """
    try:
        if value == None:
            return None
        # Use regex to extract the numeric part and unit
        match = re.match(r"([\d,\.]+)([BMK]?)", value.strip(), re.IGNORECASE)

        if not match:
            return 0.0  # Return 0 if value is not in expected format

        num_part, unit = match.groups()
        num_part = float(num_part.replace(",", ""))  # Convert numeric part to float

        # Apply conversion based on unit
        multiplier = {"B": 1_000_000_000, "M": 1_000_000, "K": 1_000}
        return num_part * multiplier.get(unit.upper(), 1)  # Default to 1 if no unit

    except Exception as e:
        print(f"Error converting value: {value}, Error: {e}")
        return 0.0  # Return 0.0 in case of errors


def scrape_yahoo_summary(ticker):
    """Scrapes summary page details from Yahoo Finance."""
    url = f"https://finance.yahoo.com/quote/{ticker}?p={ticker}"
    headers = {"User-Agent": "Mozilla/5.0"}

    try:
        response = requests.get(url, headers=headers)
        soup = BeautifulSoup(response.text, "html.parser")

        summary_data = {}
        for row in soup.find_all("li"):
            cols = row.find_all("p")
            if len(cols) >= 2:
                key = cols[0].text.strip()
                value = cols[1].text.strip()
                summary_data[key] = value

        return summary_data
    except Exception as e:
        print(f"Error scraping summary for {ticker}: {e}")
        return {}

def getSummarizedData(stock_data,):
    try:
        symbol = stock_data.get("info", {}).get("symbol")
        financials = stock_data.get("financials", {})
        yearly_income_statement = financials.get("yearly_income_statement", {})

        # Extract "Total Revenue" row
        total_revenue = yearly_income_statement.loc["Total Revenue"] if "Total Revenue" in yearly_income_statement.index else None

        if total_revenue is None or total_revenue.empty:
            print("Warning: Total Revenue data not found.")
            return {key: 0.0 for key in [
                "Rev Growth 5 years CAGR", "Rev Growth 3 years CAGR", "Rev Growth 1 year",
                "Gross Margin (%)", "Operating Margin (%)", "Net Profit Margin (%)",
                "Debt/Equity", "FCF/Net Profit","PRICE/BOOK","ROA"
            ]}

        total_revenue = total_revenue.sort_index(ascending=False)  # Ensure latest year is first
        latest_year = total_revenue.iloc[0] if len(total_revenue) > 0 else 0.0

        # Extract historical revenue, return 0 if missing
        revenue_1y_back = total_revenue.iloc[1] if len(total_revenue) > 1 else 0.0
        revenue_3y_back = total_revenue.iloc[3] if len(total_revenue) > 3 else 0.0
        revenue_5y_back = total_revenue.iloc[5] if len(total_revenue) > 5 else 0.0

        # Revenue Growth Calculations (Avoid ZeroDivisionError)
        try:
            revGrowth_5Year = pow((latest_year / revenue_5y_back), 1 / 5) - 1 if revenue_5y_back > 0 else 0.0
            revGrowth_3Year = pow((latest_year / revenue_3y_back), 1 / 3) - 1 if revenue_3y_back > 0 else 0.0
            revGrowth_1Year = (latest_year / revenue_1y_back) - 1 if revenue_1y_back > 0 else 0.0
        except ZeroDivisionError:
            revGrowth_5Year, revGrowth_3Year, revGrowth_1Year = 0.0, 0.0, 0.0

        # Get financial ratios, return 0.0 if missing
        stock_info = stock_data.get("info", {})
        grossMargins = stock_info.get("grossMargins", 0.0) * 100
        operatingMargins = stock_info.get("operatingMargins", 0.0) * 100
        profitMargins = stock_info.get("profitMargins", 0.0) * 100
        debtToEquity = stock_info.get("debtToEquity", 0.0)
        if profitMargins == 0 :
            statistics = scrape_yahoo_statistics(symbol)
            try:
                profitMargins = convert_to_numeric(statistics.get("Net Income Avi to Common  (ttm)")) / convert_to_numeric(statistics.get("Revenue  (ttm)"))
            except ZeroDivisionError:
                profitMargins=0.0

        summary = scrape_yahoo_summary(symbol)
        price_book = convert_to_numeric(summary.get("Price/Book  (mrq)"))
        roa =summary.get("Return on Assets  (ttm)")

        # Free Cash Flow to Net Profit Ratio (Avoid ZeroDivisionError)
        operatingCashflow = stock_info.get("operatingCashflow", 0.0)
        netIncomeToCommon = stock_info.get("netIncomeToCommon", 1)  # Default to 1 to avoid division by zero
        try:
            fcfNetProfit = operatingCashflow / netIncomeToCommon if netIncomeToCommon != 0 else 0.0
        except ZeroDivisionError:
            fcfNetProfit = 0.0
        return {
            "Rev Growth 5 years CAGR": round(revGrowth_5Year, 2),
            "Rev Growth 3 years CAGR": round(revGrowth_3Year, 2),
            "Rev Growth 1 year": round(revGrowth_1Year, 2),
            "Gross Margin (%)": round(grossMargins, 2),
            "Operating Margin (%)": round(operatingMargins, 2),
            "Net Profit Margin (%)": round(profitMargins, 2),
            "Debt/Equity": round(debtToEquity, 2),
            "FCF/Net Profit": round(fcfNetProfit, 2),
            "PRICE/BOOK": price_book,
            "ROA": roa
        }

    except Exception as e:
        print(f"Error fetching summarized data: {e}")
        return {key: 0.0 for key in [
            "Rev Growth 5 years CAGR", "Rev Growth 3 years CAGR", "Rev Growth 1 year",
            "Gross Margin (%)", "Operating Margin (%)", "Net Profit Margin (%)",
            "Debt/Equity", "FCF/Net Profit", "PRICE/BOOK","ROA"
        ]}



def update_stock_data_in_excel(original_file, updated_file):
    """
    Reads stock symbols from the first column of the Excel file, fetches financial data,
    and updates respective columns in a duplicate file while preserving formatting.

    :param original_file: Path to the original Excel file.
    :param updated_file: Path where the modified Excel file will be saved.
    """
    try:
        # Create a duplicate of the original file to preserve the original
        shutil.copy(original_file, updated_file)

        # Load the duplicated Excel file with openpyxl
        wb = openpyxl.load_workbook(updated_file)
        sheet = wb.active  # Use the first sheet

        # Identify the first column as the symbol column
        symbol_column = 1  # Assuming the first column has stock symbols

        # Read header row (column names)
        headers = {cell.value: cell.column for cell in sheet[1] if cell.value}
        print("headers == ",headers)
        # Iterate over each row to fetch and update stock data
        counter =0;
        for row in range(2, sheet.max_row + 1):  # Skip header row
            # if counter>3:
            #     break
            stock_symbol = str(sheet.cell(row=row, column=symbol_column).value).strip()
            counter+=1
            if not stock_symbol or stock_symbol.lower() == "none":
                continue  # Skip empty rows

            try:
                # Fetch stock data
                stock_data = YahooFinanceService().fetch_stock_data(stock_symbol)
                summary_data = getSummarizedData(stock_data)  # Get summarized financials
                # Update respective columns with the fetched data
                for key, value in summary_data.items():
                    if key in headers:  # Update only if the column exists
                        sheet.cell(row=row, column=headers[key], value=value)
                    else:
                        print("Not found ",key)
                print(f"Updated data for {stock_symbol}")

            except Exception as e:
                print(f"Error processing {stock_symbol}: {e}")

        # Save changes to the duplicate file
        wb.save(updated_file)
        print(f"Updated data saved to: {updated_file}")

    except Exception as e:
        print(f"Error updating Excel file: {e}")

def main():
    stock_symbols = read_stock_symbols()
    if not stock_symbols:
        print("No stock symbols to process.")
        return

    stock_summaries = []

    for stock in stock_symbols[:]:  # Create a copy to modify list safely
        try:
            print(f"Processing stock: {stock}")

            stock_data = YahooFinanceService().fetch_stock_data(stock)

            # Remove stock from list & update file
            # stock_symbols.remove(stock)
            update_stock_file(stock_symbols)
            data=  getSummarizedData(stock_data)
            # Append to completed file
            append_to_completed(stock)

            # Save Excel file after every iteration
            print(f"Excel report updated: {EXCEL_FILE}")
        except Exception as e:
            print(f"Error processing {stock}: {e}")
            traceback.print_exc()  # Print full stack trace

            # Append the failed symbol to a failed file
            append_to_failed(stock)

            # Remove the failed symbol from the list and update the stock file
            # stock_symbols.remove(stock)
            # update_stock_file(stock_symbols)

            # Continue processing the next symbol
            continue

    print("Processing complete!")


if __name__ == "__main__":
    # main()
    original_file="C:\\Users\\prsh01\\Downloads\\StockAnalyzer\\Assessment.xlsx"
    updated_file=f"{getOutputDirectory()}\\Assessment.xlsx"
    update_stock_data_in_excel(original_file,updated_file)
